import pyautogui
from pynput import mouse
from openpyxl import Workbook, load_workbook
import os

# Nombre del archivo Excel
archivo_excel = "coordenadas.xlsx"

# Crear Excel si no existe
if not os.path.exists(archivo_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clicks"
    ws.append(["N°", "X", "Y"])  # Encabezados
    wb.save(archivo_excel)

# Cargar archivo existente
wb = load_workbook(archivo_excel)
ws = wb.active

contador = ws.max_row  # Para seguir numerando

print("📍 Programa iniciado. Haz clic en cualquier parte de la pantalla.")
print("❌ Presiona ESC para salir.")

def on_click(x, y, button, pressed):
    global contador
    if pressed:  # Solo al presionar (no al soltar)
        contador += 1
        print(f"🖱️ Click #{contador-1} en coordenadas: {x}, {y}")
        ws.append([contador-1, x, y])  # Guardar en Excel
        wb.save(archivo_excel)

# Iniciar escucha de mouse
with mouse.Listener(on_click=on_click) as listener:
    listener.join()
